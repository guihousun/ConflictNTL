from __future__ import annotations

import math
import os
from pathlib import Path

import ee
import geopandas as gpd
import pandas as pd
from dotenv import load_dotenv
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pyproj import CRS
from shapely.geometry import mapping
from shapely.ops import transform, unary_union
from pyproj import Transformer


DATA = Path(r"D:\Research_vault\raw\writing\conflictntl\data")
GEOB_ADMIN_DIR = Path(r"D:\Research_vault\raw\datasets\geoboundaries")
EVENTS = DATA / "event_screening_geoboundaries_v2_qgis" / "events_osm_v2_downstream.csv"
OLD_CLUSTERS = DATA / "event_screening_geoboundaries" / "buffer_clusters_5km.csv"
OLD_METRICS = DATA / "event_screening_geoboundaries" / "buffer_ntl_v2" / "v2_buffer_5km_change_metrics.csv"
OUT_DIR = DATA / "event_screening_geoboundaries_v2_qgis" / "buffer_ntl_aeqd5"
OUT_XLSX = (
    Path(r"D:\Research_vault\raw\writing\conflictntl\attachments")
    / "table1_aeqd_5km_cluster_ntl_complete.xlsx"
)

ANALYSIS_START = "2026-02-13"
PREWAR_END = "2026-02-26"
CONFLICT_START = "2026-02-27"
CONFLICT_END = "2026-04-07"
CEASEFIRE_START = "2026-04-08"
ANALYSIS_END = "2026-04-21"
EE_END_EXCLUSIVE = "2026-04-22"
BAND = "DNB_BRDF_Corrected_NTL"
DATASET = "NASA/VIIRS/002/VNP46A2"
SCALE_M = 500
RADIUS_KM = 5
ADM2_CACHE: gpd.GeoDataFrame | None = None

PERIODS = [
    ("T1_prewar", ANALYSIS_START, PREWAR_END),
    ("T2_conflict", CONFLICT_START, CONFLICT_END),
    ("T3_ceasefire", CEASEFIRE_START, ANALYSIS_END),
]


class UnionFind:
    def __init__(self, n: int) -> None:
        self.parent = list(range(n))
        self.rank = [0] * n

    def find(self, x: int) -> int:
        while self.parent[x] != x:
            self.parent[x] = self.parent[self.parent[x]]
            x = self.parent[x]
        return x

    def union(self, a: int, b: int) -> None:
        ra, rb = self.find(a), self.find(b)
        if ra == rb:
            return
        if self.rank[ra] < self.rank[rb]:
            self.parent[ra] = rb
        elif self.rank[ra] > self.rank[rb]:
            self.parent[rb] = ra
        else:
            self.parent[rb] = ra
            self.rank[ra] += 1


def clean_mode(values: pd.Series) -> str:
    cleaned = []
    for value in values.dropna().astype(str):
        v = value.strip()
        if v and v.lower() not in {"nan", "unknown", "unk", "na", "n/a"}:
            cleaned.append(v)
    if not cleaned:
        return ""
    return pd.Series(cleaned).mode().iloc[0]


def load_adm2_boundaries() -> gpd.GeoDataFrame:
    global ADM2_CACHE
    if ADM2_CACHE is not None:
        return ADM2_CACHE

    frames = []
    for iso3 in ["IRN", "ISR"]:
        path = GEOB_ADMIN_DIR / iso3 / f"osm_gee_geoboundaries_{iso3}_adm2.geojson"
        adm2 = gpd.read_file(path)[["shapeName", "iso3", "geometry"]].to_crs("EPSG:4326")
        adm2["shapeName"] = adm2["shapeName"].astype(str).str.strip()
        frames.append(adm2)
    ADM2_CACHE = gpd.GeoDataFrame(pd.concat(frames, ignore_index=True), crs="EPSG:4326")
    return ADM2_CACHE


def adm2_area_name(cluster_geom_wgs84, country: str) -> str:
    adm2 = load_adm2_boundaries()
    iso3 = {"Iran": "IRN", "Israel": "ISR"}.get(country)
    scoped = adm2[adm2["iso3"].eq(iso3)] if iso3 else adm2
    scoped = scoped[scoped.geometry.intersects(cluster_geom_wgs84)].copy()
    if scoped.empty:
        return ""

    geom_equal_area = gpd.GeoSeries([cluster_geom_wgs84], crs="EPSG:4326").to_crs("EPSG:6933").iloc[0]
    scoped_equal_area = scoped.to_crs("EPSG:6933")
    scored = []
    for original, equal_area_geom in zip(scoped.itertuples(index=False), scoped_equal_area.geometry):
        area = equal_area_geom.intersection(geom_equal_area).area
        if area > 0:
            scored.append((str(original.shapeName).strip(), float(area)))
    if not scored:
        scored = [(str(row.shapeName).strip(), 0.0) for row in scoped.itertuples(index=False)]

    ordered = []
    seen = set()
    for name, _area in sorted(scored, key=lambda item: (-item[1], item[0])):
        if name and name.lower() not in {"nan", "unknown"} and name not in seen:
            ordered.append(name)
            seen.add(name)
    return "&".join(ordered)


def load_events() -> gpd.GeoDataFrame:
    events = pd.read_csv(EVENTS, encoding="utf-8-sig")
    events["source_row_id"] = range(1, len(events) + 1)
    events["latitude"] = pd.to_numeric(events["latitude"], errors="coerce")
    events["longitude"] = pd.to_numeric(events["longitude"], errors="coerce")
    events["event_id"] = events["event_id"].astype(str)
    events = events.dropna(subset=["latitude", "longitude"]).reset_index(drop=True)
    return gpd.GeoDataFrame(
        events,
        geometry=gpd.points_from_xy(events["longitude"], events["latitude"]),
        crs="EPSG:4326",
    )


def build_aeqd_clusters(events_gdf: gpd.GeoDataFrame):
    lon0 = float(events_gdf["longitude"].mean())
    lat0 = float(events_gdf["latitude"].mean())
    aeqd = CRS.from_proj4(
        f"+proj=aeqd +lat_0={lat0:.8f} +lon_0={lon0:.8f} +datum=WGS84 +units=m +no_defs"
    )
    metric = events_gdf.to_crs(aeqd).reset_index(drop=True)
    sindex = metric.sindex
    uf = UnionFind(len(metric))
    threshold_m = RADIUS_KM * 2 * 1000

    for i, geom in enumerate(metric.geometry):
        for j in sindex.query(geom.buffer(threshold_m), predicate="intersects"):
            j = int(j)
            if j <= i:
                continue
            if geom.distance(metric.geometry.iloc[j]) <= threshold_m:
                uf.union(i, j)

    roots = [uf.find(i) for i in range(len(metric))]
    root_sizes = pd.Series(roots).value_counts()
    root_order = root_sizes.sort_values(ascending=False).index.tolist()
    root_to_rank = {root: rank for rank, root in enumerate(root_order, start=1)}
    metric["cluster_root"] = roots
    metric["rank"] = metric["cluster_root"].map(root_to_rank)
    metric["aeqd_cluster_id"] = metric["rank"].map(lambda x: f"AEQD5_{int(x):03d}")

    to_wgs84 = Transformer.from_crs(aeqd, "EPSG:4326", always_xy=True).transform

    cluster_rows = []
    cluster_geoms_wgs84 = []
    for rank, group in metric.groupby("rank", sort=True):
        original = events_gdf.iloc[group.index].copy()
        buffers = list(group.geometry.buffer(RADIUS_KM * 1000))
        cluster_geom_metric = unary_union(buffers)
        if not cluster_geom_metric.is_valid:
            cluster_geom_metric = cluster_geom_metric.buffer(0)
        cluster_geom_wgs84 = transform(to_wgs84, cluster_geom_metric)
        cluster_geoms_wgs84.append(cluster_geom_wgs84)

        conflict_count = int(original["date"].between(CONFLICT_START, CONFLICT_END, inclusive="both").sum())
        ceasefire_count = int(original["date"].between(CEASEFIRE_START, ANALYSIS_END, inclusive="both").sum())
        country_mode = clean_mode(original["country"])
        adm2_name = adm2_area_name(cluster_geom_wgs84, country_mode)
        cluster_rows.append(
            {
                "rank": int(rank),
                "aeqd_cluster_id": f"AEQD5_{int(rank):03d}",
                "country_mode": country_mode,
                "area_name": adm2_name or clean_mode(original["city"]) or country_mode,
                "city_mode": clean_mode(original["city"]),
                "site_subtype_mode": clean_mode(original["site_subtype"]),
                "event_type_mode": clean_mode(original["event_type"]),
                "n_event_points": int(len(original)),
                "n_unique_event_id": int(original["event_id"].nunique()),
                "duplicate_event_rows": int(len(original) - original["event_id"].nunique()),
                "T2_conflict_event_count": conflict_count,
                "T3_ceasefire_event_count": ceasefire_count,
                "first_event_date": str(original["date"].min()),
                "last_event_date": str(original["date"].max()),
                "centroid_lon": float(original["longitude"].mean()),
                "centroid_lat": float(original["latitude"].mean()),
                "exact_coord_share": float(original["coord_type"].astype(str).str.lower().eq("exact").mean()),
                "buffer_union_area_km2_aeqd": float(cluster_geom_metric.area / 1_000_000),
            }
        )

    clusters = pd.DataFrame(cluster_rows).sort_values("rank").reset_index(drop=True)
    clusters_gdf = gpd.GeoDataFrame(clusters, geometry=cluster_geoms_wgs84, crs="EPSG:4326")
    membership = metric.drop(columns="geometry").copy()
    membership = membership[
        [
            "aeqd_cluster_id",
            "rank",
            "cluster_root",
            "source_row_id",
            "event_id",
            "date",
            "country",
            "city",
            "latitude",
            "longitude",
            "event_type",
            "site_type",
            "site_subtype",
            "coord_type",
            "llm_label",
            "building_pass",
            "building_count_5km",
        ]
    ].sort_values(["rank", "source_row_id"])
    return clusters_gdf, membership, lon0, lat0


def ee_geometry(geom):
    simplified = geom.simplify(0.00025, preserve_topology=True)
    return ee.Geometry(mapping(simplified), proj="EPSG:4326", geodesic=False)


def fetch_daily_antl(top_clusters: gpd.GeoDataFrame) -> pd.DataFrame:
    features = []
    for _, row in top_clusters.iterrows():
        features.append(
            ee.Feature(
                ee_geometry(row.geometry),
                {
                    "rank": int(row["rank"]),
                    "aeqd_cluster_id": str(row["aeqd_cluster_id"]),
                    "area_name": str(row["area_name"]),
                },
            )
        )
    fc = ee.FeatureCollection(features)
    image_collection = ee.ImageCollection(DATASET).filterDate(ANALYSIS_START, EE_END_EXCLUSIVE).select(BAND)
    projection = ee.Image(image_collection.first()).projection()

    total_pixel_features = (
        ee.Image.constant(1)
        .rename("total_pixels")
        .reproject(projection)
        .reduceRegions(collection=fc, reducer=ee.Reducer.count(), scale=SCALE_M, tileScale=4)
        .getInfo()
    )
    total_pixels = {}
    for f in total_pixel_features.get("features", []):
        props = f.get("properties", {})
        total_pixels[str(props.get("aeqd_cluster_id"))] = props.get("count", math.nan)

    def reduce_image(img):
        date = img.date().format("YYYY-MM-dd")
        reduced = img.select(BAND).reduceRegions(
            collection=fc,
            reducer=ee.Reducer.mean().combine(reducer2=ee.Reducer.count(), sharedInputs=True),
            scale=SCALE_M,
            tileScale=4,
        )
        return reduced.map(lambda f: f.set("date", date))

    collection = image_collection.map(reduce_image).flatten()
    info = collection.getInfo()
    rows = []
    for f in info.get("features", []):
        props = f.get("properties", {})
        cluster_id = str(props.get("aeqd_cluster_id"))
        rows.append(
            {
                "rank": int(props.get("rank")),
                "aeqd_cluster_id": cluster_id,
                "area_name": props.get("area_name"),
                "date": props.get("date"),
                "ANTL": props.get("mean", math.nan),
                "valid_pixel_count": props.get("count", math.nan),
                "total_pixel_count": total_pixels.get(cluster_id, math.nan),
            }
        )
    out = pd.DataFrame(rows)
    if out.empty:
        return out
    out["ANTL"] = pd.to_numeric(out["ANTL"], errors="coerce")
    out["valid_pixel_count"] = pd.to_numeric(out["valid_pixel_count"], errors="coerce")
    out["total_pixel_count"] = pd.to_numeric(out["total_pixel_count"], errors="coerce")
    max_valid = out.groupby("aeqd_cluster_id")["valid_pixel_count"].transform("max")
    out["total_pixel_count"] = out[["total_pixel_count"]].join(max_valid.rename("max_valid")).max(axis=1)
    out["valid_pixel_ratio"] = (out["valid_pixel_count"] / out["total_pixel_count"]).clip(upper=1)
    return out.sort_values(["rank", "date"]).reset_index(drop=True)


def build_daily_panel(top_clusters: pd.DataFrame, membership: pd.DataFrame, antl_df: pd.DataFrame) -> pd.DataFrame:
    dates = pd.date_range(ANALYSIS_START, ANALYSIS_END, freq="D").strftime("%Y-%m-%d")
    base = top_clusters[["rank", "aeqd_cluster_id", "area_name"]].merge(pd.DataFrame({"date": dates}), how="cross")
    panel = base.merge(antl_df, on=["rank", "aeqd_cluster_id", "area_name", "date"], how="left")
    event_counts = (
        membership[membership["aeqd_cluster_id"].isin(top_clusters["aeqd_cluster_id"])]
        .groupby(["aeqd_cluster_id", "date"])
        .size()
        .reset_index(name="event_count")
    )
    panel = panel.merge(event_counts, on=["aeqd_cluster_id", "date"], how="left")
    panel["event_count"] = panel["event_count"].fillna(0).astype(int)
    return panel


def summarize_periods(panel: pd.DataFrame, top_clusters: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for _, cluster in top_clusters.iterrows():
        sub = panel[panel["aeqd_cluster_id"] == cluster["aeqd_cluster_id"]].copy()
        for period, start, end in PERIODS:
            mask = (sub["date"] >= start) & (sub["date"] <= end)
            scoped = sub.loc[mask].copy()
            rows.append(
                {
                    "rank": int(cluster["rank"]),
                    "aeqd_cluster_id": str(cluster["aeqd_cluster_id"]),
                    "area_name": str(cluster["area_name"]),
                    "period": period,
                    "mean_ANTL": float(pd.to_numeric(scoped["ANTL"], errors="coerce").mean()),
                    "valid_day_count": int(pd.to_numeric(scoped["ANTL"], errors="coerce").notna().sum()),
                    "mean_valid_pixel_ratio": float(pd.to_numeric(scoped["valid_pixel_ratio"], errors="coerce").mean()),
                    "period_event_count": int(scoped["event_count"].sum()),
                }
            )
    return pd.DataFrame(rows)


def build_change_metrics(period_df: pd.DataFrame, top_clusters: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for _, cluster in top_clusters.iterrows():
        sub = period_df[period_df["aeqd_cluster_id"] == cluster["aeqd_cluster_id"]].set_index("period")
        t1 = float(sub.loc["T1_prewar", "mean_ANTL"])
        t2 = float(sub.loc["T2_conflict", "mean_ANTL"])
        t3 = float(sub.loc["T3_ceasefire", "mean_ANTL"])
        row = cluster.drop(labels=["geometry"]).to_dict()
        row.update(
            {
                "T1_prewar_ANTL": t1,
                "T2_conflict_ANTL": t2,
                "T3_ceasefire_ANTL": t3,
                "T2_vs_T1_pct": (t2 - t1) / t1 * 100.0 if t1 else math.nan,
                "T3_vs_T2_pct": (t3 - t2) / t2 * 100.0 if t2 else math.nan,
                "T1_valid_day_count": int(sub.loc["T1_prewar", "valid_day_count"]),
                "T2_valid_day_count": int(sub.loc["T2_conflict", "valid_day_count"]),
                "T3_valid_day_count": int(sub.loc["T3_ceasefire", "valid_day_count"]),
                "T1_mean_valid_pixel_ratio": float(sub.loc["T1_prewar", "mean_valid_pixel_ratio"]),
                "T2_mean_valid_pixel_ratio": float(sub.loc["T2_conflict", "mean_valid_pixel_ratio"]),
                "T3_mean_valid_pixel_ratio": float(sub.loc["T3_ceasefire", "mean_valid_pixel_ratio"]),
            }
        )
        rows.append(row)
    return pd.DataFrame(rows)


def write_excel(table1: pd.DataFrame, clusters: pd.DataFrame, membership: pd.DataFrame, daily: pd.DataFrame, period: pd.DataFrame, lon0: float, lat0: float) -> None:
    old_clusters = pd.read_csv(OLD_CLUSTERS, encoding="utf-8-sig")
    old_metrics = pd.read_csv(OLD_METRICS, encoding="utf-8-sig")
    old_top10 = (
        old_metrics.sort_values("rank")
        .head(10)
        .merge(
            old_clusters[["cluster_id", "event_count", "lat", "lon"]],
            on="cluster_id",
            how="left",
            suffixes=("_metrics", "_cluster_table"),
        )
    )
    validation = pd.DataFrame(
        [
            ["source_event_rows", len(membership)],
            ["source_unique_event_id", int(membership["event_id"].nunique())],
            ["aeqd_center_lon0", lon0],
            ["aeqd_center_lat0", lat0],
            ["aeqd_5km_cluster_count", len(clusters)],
            ["top10_daily_rows", len(daily)],
            ["gee_dataset", DATASET],
            ["gee_band", BAND],
            ["gee_scale_m", SCALE_M],
            ["analysis_window", f"{ANALYSIS_START} to {ANALYSIS_END}; EE end-exclusive {EE_END_EXCLUSIVE}"],
            ["distance_rule", "AEQD point distance <= 10 km; equivalent to intersecting 5 km buffers"],
            ["main_error_source", "Old QGIS experiment used non-equidistant metric CRS or lon/lat degree buffers for later cluster/NTL steps, so event counts, AOI geometry, and NTL statistics were not one consistent object."],
        ],
        columns=["item", "value"],
    )

    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        table1.to_excel(writer, sheet_name="Table1_complete", index=False)
        clusters.drop(columns=["geometry"], errors="ignore").to_excel(writer, sheet_name="All_184_clusters", index=False)
        period.to_excel(writer, sheet_name="Top10_period_summary", index=False)
        daily.to_excel(writer, sheet_name="Top10_daily_ANTL", index=False)
        membership.to_excel(writer, sheet_name="Membership_2383_points", index=False)
        old_top10.to_excel(writer, sheet_name="Old_308_top10_compare", index=False)
        validation.to_excel(writer, sheet_name="Validation_notes", index=False)

        wb = writer.book
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            ws.sheet_view.showGridLines = False
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for col_idx, col_cells in enumerate(ws.columns, start=1):
                max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells)
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 38)
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=False)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    events = load_events()
    clusters_gdf, membership, lon0, lat0 = build_aeqd_clusters(events)
    top10 = clusters_gdf.head(10).copy()
    clusters_gdf.to_file(OUT_DIR / "aeqd_5km_clusters_all_184.geojson", driver="GeoJSON")
    top10.to_file(OUT_DIR / "aeqd_5km_top10_clusters.geojson", driver="GeoJSON")
    membership.to_csv(OUT_DIR / "aeqd_5km_membership_2383_points.csv", index=False, encoding="utf-8-sig")
    clusters_gdf.drop(columns="geometry").to_csv(OUT_DIR / "aeqd_5km_clusters_all_184.csv", index=False, encoding="utf-8-sig")

    load_dotenv(Path(r"D:\NTL-GPT-Clone") / ".env")
    ee.Initialize(project=os.getenv("GEE_DEFAULT_PROJECT_ID"))
    daily = fetch_daily_antl(top10)
    daily.to_csv(OUT_DIR / "aeqd_5km_top10_daily_antl.csv", index=False, encoding="utf-8-sig")
    panel = build_daily_panel(top10.drop(columns="geometry"), membership, daily)
    panel.to_csv(OUT_DIR / "aeqd_5km_top10_daily_panel.csv", index=False, encoding="utf-8-sig")
    period = summarize_periods(panel, top10.drop(columns="geometry"))
    period.to_csv(OUT_DIR / "aeqd_5km_top10_period_summary.csv", index=False, encoding="utf-8-sig")
    table1 = build_change_metrics(period, top10)
    table1.to_csv(OUT_DIR / "aeqd_5km_top10_table1_complete.csv", index=False, encoding="utf-8-sig")
    write_excel(table1, clusters_gdf, membership, daily, period, lon0, lat0)

    print(f"Wrote {OUT_XLSX}")
    print(table1[["rank", "aeqd_cluster_id", "area_name", "n_event_points", "T1_prewar_ANTL", "T2_conflict_ANTL", "T3_ceasefire_ANTL", "T2_vs_T1_pct", "T3_vs_T2_pct"]].to_string(index=False))


if __name__ == "__main__":
    main()
