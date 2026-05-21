from __future__ import annotations

import importlib.util
import math
import os
from pathlib import Path

import ee
import pandas as pd
from dotenv import load_dotenv
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HERE = Path(__file__).resolve().parent
BASE_SCRIPT = HERE / "make_table1_aeqd_5km_clusters_ntl_complete.py"
DATA = Path(r"D:\Research_vault\raw\writing\conflictntl\data")
OUT_DIR = DATA / "event_screening_geoboundaries_v2_qgis" / "buffer_ntl_aeqd3"
OUT_XLSX = (
    Path(r"D:\Research_vault\raw\writing\conflictntl\attachments")
    / "table1_aeqd_3km_cluster_ntl_complete.xlsx"
)


def load_base_module():
    spec = importlib.util.spec_from_file_location("aeqd5_table_base", BASE_SCRIPT)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Cannot load {BASE_SCRIPT}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    module.RADIUS_KM = 3
    module.OUT_DIR = OUT_DIR
    module.OUT_XLSX = OUT_XLSX
    return module


def relabel_to_3km(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for col in out.columns:
        if col == "aeqd_cluster_id":
            out[col] = out[col].astype(str).str.replace("AEQD5_", "AEQD3_", regex=False)
    return out


def write_excel_3km(
    base,
    table1: pd.DataFrame,
    clusters,
    membership: pd.DataFrame,
    daily: pd.DataFrame,
    period: pd.DataFrame,
    lon0: float,
    lat0: float,
) -> None:
    validation = pd.DataFrame(
        [
            ["source_event_rows", len(membership)],
            ["source_unique_event_id", int(membership["event_id"].nunique())],
            ["aeqd_center_lon0", lon0],
            ["aeqd_center_lat0", lat0],
            ["aeqd_3km_cluster_count", len(clusters)],
            ["top10_daily_rows", len(daily)],
            ["gee_dataset", base.DATASET],
            ["gee_band", base.BAND],
            ["gee_scale_m", base.SCALE_M],
            ["analysis_window", f"{base.ANALYSIS_START} to {base.ANALYSIS_END}; EE end-exclusive {base.EE_END_EXCLUSIVE}"],
            ["distance_rule", "AEQD point distance <= 6 km; equivalent to intersecting 3 km buffers"],
            ["derivation", f"3 km sensitivity version derived from {BASE_SCRIPT.name}; source events are unchanged 2383 downstream points."],
        ],
        columns=["item", "value"],
    )

    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        table1.to_excel(writer, sheet_name="Table1_complete", index=False)
        clusters.drop(columns=["geometry"], errors="ignore").to_excel(writer, sheet_name=f"All_{len(clusters)}_clusters", index=False)
        period.to_excel(writer, sheet_name="Top10_period_summary", index=False)
        daily.to_excel(writer, sheet_name="Top10_daily_ANTL", index=False)
        membership.to_excel(writer, sheet_name="Membership_2383_points", index=False)
        validation.to_excel(writer, sheet_name="Validation_notes", index=False)

        wb = writer.book
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            ws.sheet_view.showGridLines = False
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for col_idx, col_cells in enumerate(ws.columns, start=1):
                max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells)
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 38)
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=False)


def main() -> None:
    base = load_base_module()
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    events = base.load_events()
    clusters_gdf, membership, lon0, lat0 = base.build_aeqd_clusters(events)
    clusters_gdf = relabel_to_3km(clusters_gdf)
    membership = relabel_to_3km(membership)
    top10 = clusters_gdf.head(10).copy()

    cluster_count = len(clusters_gdf)
    clusters_gdf.to_file(OUT_DIR / f"aeqd_3km_clusters_all_{cluster_count}.geojson", driver="GeoJSON")
    top10.to_file(OUT_DIR / "aeqd_3km_top10_clusters.geojson", driver="GeoJSON")
    membership.to_csv(OUT_DIR / "aeqd_3km_membership_2383_points.csv", index=False, encoding="utf-8-sig")
    clusters_gdf.drop(columns="geometry").to_csv(
        OUT_DIR / f"aeqd_3km_clusters_all_{cluster_count}.csv",
        index=False,
        encoding="utf-8-sig",
    )

    load_dotenv(Path(r"D:\NTL-GPT-Clone") / ".env")
    ee.Initialize(project=os.getenv("GEE_DEFAULT_PROJECT_ID"))
    daily = base.fetch_daily_antl(top10)
    daily = relabel_to_3km(daily)
    daily.to_csv(OUT_DIR / "aeqd_3km_top10_daily_antl.csv", index=False, encoding="utf-8-sig")
    panel = base.build_daily_panel(top10.drop(columns="geometry"), membership, daily)
    panel = relabel_to_3km(panel)
    panel.to_csv(OUT_DIR / "aeqd_3km_top10_daily_panel.csv", index=False, encoding="utf-8-sig")
    period = base.summarize_periods(panel, top10.drop(columns="geometry"))
    period = relabel_to_3km(period)
    period.to_csv(OUT_DIR / "aeqd_3km_top10_period_summary.csv", index=False, encoding="utf-8-sig")
    table1 = base.build_change_metrics(period, top10)
    table1 = relabel_to_3km(table1)
    table1.to_csv(OUT_DIR / "aeqd_3km_top10_table1_complete.csv", index=False, encoding="utf-8-sig")
    write_excel_3km(base, table1, clusters_gdf, membership, daily, period, lon0, lat0)

    print(f"Wrote {OUT_XLSX}")
    print(f"AEQD 3 km cluster count: {cluster_count}")
    cols = [
        "rank",
        "aeqd_cluster_id",
        "area_name",
        "n_event_points",
        "T1_prewar_ANTL",
        "T2_conflict_ANTL",
        "T3_ceasefire_ANTL",
        "T2_vs_T1_pct",
        "T3_vs_T2_pct",
    ]
    print(table1[cols].to_string(index=False))

    if table1[["T1_prewar_ANTL", "T2_conflict_ANTL", "T3_ceasefire_ANTL"]].isna().any().any():
        raise RuntimeError("NaN period ANTL detected in AEQD 3 km Table 1 outputs.")
    if not math.isclose(float(len(membership)), 2383.0):
        raise RuntimeError(f"Unexpected membership row count: {len(membership)}")


if __name__ == "__main__":
    main()
